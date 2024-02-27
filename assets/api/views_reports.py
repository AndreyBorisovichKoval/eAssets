# pip install pandas
# pip install openpyxl
# pip install xlsxwriter

import pandas as pd
import openpyxl as xl
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.drawing.image import Image
# from openpyxl.styles import Alignment
from django.utils import timezone
from datetime import datetime, timezone
from django.db.models import Sum
from django.shortcuts import get_object_or_404
from django.http import JsonResponse, HttpResponse
from rest_framework import status, viewsets, request
from rest_framework.exceptions import AuthenticationFailed
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated
# from rest_framework_simplejwt.models.TokenUser import username
from rest_framework_simplejwt.tokens import AccessToken
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from typing import Any
from assets.api.serializers import (UserSerializer, DepartmentSerializer, DivisionSerializer, PositionSerializer,
                                    StaffSerializer, AssetTypeSerializer, AssetSerializer, AssetAssignmentSerializer)
from assets.models import *
import logging

logger = logging.getLogger('django')


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def get_staff_assets(request, pk):
    staff = get_object_or_404(Staff, pk=pk)
    assigned_assets = Asset.objects.filter(assetassignment__staff=staff)

    asset_list = []
    for asset in assigned_assets:
        asset_list.append({
            'asset_type': asset.asset_type.title if asset.asset_type else None,
            'inventory_number': asset.inventory_number,
            'title': asset.title,
            'identifier': asset.identifier,
        })

    logger.info(f"__-=*=-__ Retrieved assigned assets for staff with ID: {pk}. __-=*=-__")
    return JsonResponse(asset_list, safe=False)


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def get_department_assets(request, pk):
    department = get_object_or_404(Department, id=pk)
    staff_in_department = Staff.objects.filter(division__department=department)
    assigned_assets = Asset.objects.filter(assetassignment__staff__in=staff_in_department)

    asset_list = []
    for asset in assigned_assets:
        asset_list.append({
            'asset_type': asset.asset_type.title if asset.asset_type else None,
            'inventory_number': asset.inventory_number,
            'title': asset.title,
            'identifier': asset.identifier,
        })

    logger.info(f"Retrieved assets for department with ID: {pk}")
    return JsonResponse(asset_list, safe=False)


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def get_division_assets(request, pk):
    division = get_object_or_404(Division, id=pk)
    staff_in_division = Staff.objects.filter(division=division)
    assigned_assets = Asset.objects.filter(assetassignment__staff__in=staff_in_division)

    asset_list = []
    for asset in assigned_assets:
        asset_assignments = asset.assetassignment_set.all()
        staff_full_names = [assignment.staff.full_name for assignment in asset_assignments]

        asset_list.append({
            'division': division.title,
            'staff_full_names': staff_full_names,
            'asset_type': asset.asset_type.title if asset.asset_type else None,
            'inventory_number': asset.inventory_number,
            'title': asset.title,
            'identifier': asset.identifier,
        })

    logger.info(f"Retrieved assets for division with ID: {pk}")
    return JsonResponse(asset_list, safe=False)


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def get_retired_assets(request):
    retired_assets = Asset.objects.filter(is_written_off=True)

    asset_list = []
    for asset in retired_assets:
        asset_list.append({
            'asset_type': asset.asset_type.title if asset.asset_type else None,
            'inventory_number': asset.inventory_number,
            'title': asset.title,
            'identifier': asset.identifier,
        })

    logger.info("Retrieved retired assets")
    return JsonResponse(asset_list, safe=False)


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def calculate_cost_assets(request):
    total_price = Asset.objects.filter(is_written_off=False).aggregate(Sum('cost'))['cost__sum']
    if total_price is None:
        total_price = 0

    logger.info("Calculated total price of active assets")
    return JsonResponse({'total_price': total_price}, safe=False)


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def calculate_current_cost_assets(request):
    total_price = Asset.objects.filter(is_written_off=False).aggregate(Sum('current_cost'))['current_cost__sum']
    if total_price is None:
        total_price = 0

    logger.info("Calculated total current price of active assets")
    return JsonResponse({'total_current_price': total_price}, safe=False)


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def calculate_total_price_by_type(request, pk):
    total_price = Asset.objects.filter(asset_type_id=pk, is_written_off=False).aggregate(Sum('cost'))['cost__sum']
    if total_price is None:
        total_price = 0

    logger.info(f"Calculated total price for asset type with id {pk}")
    return JsonResponse({'total_price': total_price}, safe=False)


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def calculate_current_price_by_type(request, pk):
    current_price = Asset.objects.filter(asset_type_id=pk, is_written_off=False).aggregate(Sum('current_cost'))['current_cost__sum']
    if current_price is None:
        current_price = 0

    logger.info(f"Calculated current price for asset type with id {pk}")
    return JsonResponse({'current_price': current_price}, safe=False)



# def export_to_xlsm():
#     # Создание новой книги Excel и добавление листов
#     workbook = xl.Workbook()
#     assets_sheet = workbook.create_sheet("Assets")
#     assignments_sheet = workbook.create_sheet("Assignments")
#     positions_sheet = workbook.create_sheet("Positions")
#     divisions_sheet = workbook.create_sheet("Divisions")
#     departments_sheet = workbook.create_sheet("Departments")
#
#     # Получение данных из моделей
#     assets = Asset.objects.all()
#     assignments = AssetAssignment.objects.select_related('asset', 'staff__division__department', 'staff__position').all()
#     positions = Position.objects.all()
#     divisions = Division.objects.select_related('department').all()
#     departments = Department.objects.all()
#
#     # Заполнение листа Assets данными
#     assets_sheet.append(['Inventory Number', 'Title', 'Identifier', 'Acquisition Date', 'Service Life', 'Cost', 'Current Cost', 'Last Recalculation Date',
#                          'Description', 'Asset Type', 'Is Written Off', 'Written Off By', 'Written Off At'])
#     for asset in assets:
#         assets_sheet.append([asset.inventory_number, asset.title, asset.identifier, asset.acquisition_date, asset.service_life, asset.cost,
#                              asset.current_cost, asset.last_recalculation_date, asset.description, asset.asset_type.title,
#                              asset.is_written_off, asset.written_off_by.username if asset.written_off_by else None, asset.written_off_at])
#
#     # Заполнение листа Assignments данными
#     assignments_sheet.append(['Asset', 'Staff', 'Assignment Date', 'Return Date', 'Return By'])
#     for assignment in assignments:
#         assignments_sheet.append([assignment.asset.title, assignment.staff.full_name, assignment.assignment_date,
#                                   assignment.return_date, assignment.return_by.username if assignment.return_by else None])
#
#     # Заполнение листа Positions данными
#     positions_sheet.append(['Position'])
#     for position in positions:
#         positions_sheet.append([position.title])
#
#     # Заполнение листа Divisions данными
#     divisions_sheet.append(['Division', 'Department'])
#     for division in divisions:
#         divisions_sheet.append([division.title, division.department.title])
#
#     # Заполнение листа Departments данными
#     departments_sheet.append(['Department'])
#     for department in departments:
#         departments_sheet.append([department.title])
#
#     # Сохранение книги Excel в формате .xlsm
#     workbook.save("report.xlsm")
#
#     # Добавление макроса в файл .xlsm
#     workbook = xl.load_workbook("report.xlsm", keep_vba=True)
#     workbook.add_vba_project("path/to/vba_project.vba")
#     workbook.save("report.xlsm")
#
# export_to_xlsm()


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
# def export_to_xlsx(request):
#     assets = Asset.objects.all().values(
#         'inventory_number',
#         'title',
#         'identifier',
#         'acquisition_date',
#         'service_life',
#         'cost',
#         'current_cost',
#         'last_recalculation_date',
#         'description',
#         'asset_type',
#         'is_written_off',
#         'written_off_at'
#     )
#     assignments = AssetAssignment.objects.select_related('asset', 'staff').all().values(
#         'asset',
#         'staff',
#         'assignment_date',
#         'return_date'
#     )
#     staff = Staff.objects.select_related('division', 'position').all().values(
#         'last_name',
#         'first_name',
#         'patronymic',
#         'division__department',
#         'division__title',
#         'position'
#     )
#     divisions = Division.objects.select_related('department').all().values(
#         'department__title',
#         'title'
#     )
#     departments = Department.objects.all().values(
#         'title'
#     )
#
#     assets_df = pd.DataFrame(list(assets))
#     assignments_df = pd.DataFrame(list(assignments))
#     staff_df = pd.DataFrame(list(staff))
#     divisions_df = pd.DataFrame(list(divisions))
#     departments_df = pd.DataFrame(list(departments))
#
#     # Преобразование объектов datetime в строки без информации о часовом поясе
#     assets_df['acquisition_date'] = assets_df['acquisition_date'].apply(lambda dt: dt.strftime('%Y-%m-%d') if pd.notnull(dt) else None)
#     assignments_df['assignment_date'] = assignments_df['assignment_date'].apply(lambda dt: dt.strftime('%Y-%m-%d') if pd.notnull(dt) else None)
#     assignments_df['return_date'] = assignments_df['return_date'].apply(lambda dt: dt.strftime('%Y-%m-%d') if pd.notnull(dt) else None)
#     assets_df['last_recalculation_date'] = assets_df['last_recalculation_date'].apply(lambda dt: dt.strftime('%Y-%m-%d') if pd.notnull(dt) else None)
#     assets_df['written_off_at'] = assets_df['written_off_at'].apply(lambda dt: dt.strftime('%Y-%m-%d') if pd.notnull(dt) else None)
#
#     # Создание книги Excel с форматом .xlsx
#     with pd.ExcelWriter('report.xlsx', engine='xlsxwriter') as writer:
#         assets_df.to_excel(writer, sheet_name='Assets', index=False)
#         assignments_df.to_excel(writer, sheet_name='Assignments', index=False)
#         staff_df.to_excel(writer, sheet_name='Staff', index=False)
#         divisions_df.to_excel(writer, sheet_name='Divisions', index=False)
#         departments_df.to_excel(writer, sheet_name='Departments', index=False)
#
#     # Закрытие объекта ExcelWriter и сохранение файла .xlsx
#     writer.close()
#
#     # Отправка файла .xlsx в ответе HTTP
#     with open('report.xlsx', 'rb') as file:
#         response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=report.xlsx'
#         return response


def export_to_xlsx(request):
    assets = Asset.objects.all().values(
        'inventory_number',
        'title',
        'identifier',
        'acquisition_date',
        'service_life',
        'cost',
        'current_cost',
        'last_recalculation_date',
        'description',
        'asset_type',
        'is_written_off',
        'written_off_at'
    )
    assignments = AssetAssignment.objects.select_related('asset', 'staff').all().values(
        'asset',
        'staff',
        'assignment_date',
        'return_date'
    )
    staff = Staff.objects.select_related('division', 'position').all().values(
        'last_name',
        'first_name',
        'patronymic',
        'division__department',
        'division__title',
        'position'
    )
    divisions = Division.objects.select_related('department').all().values(
        'department__title',
        'title'
    )
    departments = Department.objects.all().values(
        'title'
    )

    assets_df = pd.DataFrame(list(assets))
    assignments_df = pd.DataFrame(list(assignments))
    staff_df = pd.DataFrame(list(staff))
    divisions_df = pd.DataFrame(list(divisions))
    departments_df = pd.DataFrame(list(departments))

    # Преобразование объектов datetime в строки без информации о часовом поясе
    assets_df['acquisition_date'] = assets_df['acquisition_date'].apply(lambda dt: dt.strftime('%Y-%m-%d') if pd.notnull(dt) else None)
    assignments_df['assignment_date'] = assignments_df['assignment_date'].apply(lambda dt: dt.strftime('%Y-%m-%d') if pd.notnull(dt) else None)
    assignments_df['return_date'] = assignments_df['return_date'].apply(lambda dt: dt.strftime('%Y-%m-%d') if pd.notnull(dt) else None)
    assets_df['last_recalculation_date'] = assets_df['last_recalculation_date'].apply(lambda dt: dt.strftime('%Y-%m-%d') if pd.notnull(dt) else None)
    assets_df['written_off_at'] = assets_df['written_off_at'].apply(lambda dt: dt.strftime('%Y-%m-%d') if pd.notnull(dt) else None)

    # Создание книги Excel с форматом .xlsx
    writer = pd.ExcelWriter('report.xlsx', engine='xlsxwriter')
    assets_df.to_excel(writer, sheet_name='Assets', index=False)
    assignments_df.to_excel(writer, sheet_name='Assignments', index=False)
    staff_df.to_excel(writer, sheet_name='Staff', index=False)
    divisions_df.to_excel(writer, sheet_name='Divisions', index=False)
    departments_df.to_excel(writer, sheet_name='Departments', index=False)

    # Закрытие объекта ExcelWriter и сохранение файла .xlsx
    writer.close()

    # Отправка файла .xlsx в ответе HTTP
    with open('report.xlsx', 'rb') as file:
        response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=report.xlsx'
        return response

